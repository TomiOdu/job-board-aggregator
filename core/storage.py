"""Reading and writing the CSV / Excel outputs."""
from __future__ import annotations

import csv
import logging
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import config
from core.models import COLUMNS, NUMERIC_COLUMNS, JobListing

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill("solid", start_color="1F3864", end_color="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF")
LINK_FONT = Font(color="0563C1", underline="single")

# Sensible max widths so the description column does not run off the screen.
COLUMN_WIDTHS = {
    "job_id": 14,
    "first_seen_date": 15,
    "last_seen_date": 15,
    "title": 40,
    "company": 28,
    "location": 26,
    "is_remote": 10,
    "salary_min": 12,
    "salary_max": 12,
    "salary_currency": 9,
    "salary_period": 12,
    "salary_is_estimate": 15,
    "salary_raw": 22,
    "contract_type": 14,
    "contract_time": 13,
    "date_posted": 13,
    "source": 10,
    "url": 46,
    "description": 70,
}


def load_master(path: Path) -> list[JobListing]:
    """Read the running dataset. Returns [] if it does not exist yet."""
    if not path.exists():
        logger.info("No existing master dataset at %s - starting fresh", path)
        return []

    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))

    listings = [JobListing.from_row(row) for row in rows]
    logger.info("Loaded %d existing listings from %s", len(listings), path.name)
    return listings


def to_dataframe(listings: list[JobListing]) -> pd.DataFrame:
    """Build a DataFrame in canonical column order with numeric salaries."""
    frame = pd.DataFrame([listing.to_row() for listing in listings], columns=COLUMNS)
    for column in NUMERIC_COLUMNS:
        frame[column] = pd.to_numeric(frame[column], errors="coerce")
    return frame


def write_csv(frame: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    # utf-8-sig: the BOM makes Windows Excel read the file as UTF-8 when you
    # double-click it, instead of cp1252 (which turns a pound sign into "Â£").
    frame.to_csv(path, index=False, encoding="utf-8-sig")
    logger.info("Wrote %d rows to %s", len(frame), path.name)


def _format_sheet(worksheet, frame: pd.DataFrame) -> None:
    """Bold header, frozen top row, autofilter, widths, clickable URLs."""
    if frame.empty:
        return

    for cell in worksheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for index, column in enumerate(frame.columns, start=1):
        letter = get_column_letter(index)
        worksheet.column_dimensions[letter].width = COLUMN_WIDTHS.get(column, 18)

        if column in NUMERIC_COLUMNS:
            for row in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row, column=index).number_format = "#,##0"

        if column == "url":
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=index)
                if cell.value:
                    cell.hyperlink = str(cell.value)
                    cell.font = LINK_FONT


def write_excel(sheets: dict[str, pd.DataFrame], path: Path) -> None:
    """Write one workbook with a sheet per entry in `sheets`."""
    path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, frame in sheets.items():
            # An empty DataFrame still needs headers written for a valid sheet.
            frame.to_excel(writer, sheet_name=sheet_name, index=False)
            _format_sheet(writer.sheets[sheet_name], frame)

    logger.info(
        "Wrote %s (%s)",
        path.name,
        ", ".join(f"{name}: {len(frame)} rows" for name, frame in sheets.items()),
    )


def write_outputs(
    full: list[JobListing], new_today: list[JobListing], run_date: date
) -> None:
    """Write every output file for this run."""
    full_frame = to_dataframe(full)
    new_frame = to_dataframe(new_today)

    write_csv(full_frame, config.MASTER_CSV)
    write_csv(new_frame, config.NEW_TODAY_CSV)
    write_excel({"All Jobs": full_frame, "New Today": new_frame}, config.LATEST_XLSX)

    if config.WRITE_ARCHIVE_COPY:
        archive_path = config.ARCHIVE_DIR / f"job_listings_{run_date.isoformat()}.xlsx"
        write_excel({"All Jobs": full_frame, "New Today": new_frame}, archive_path)
